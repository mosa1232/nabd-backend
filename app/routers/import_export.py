import io
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from .. import models
from ..database import get_db
from ..deps import require_role

router = APIRouter(prefix="/api/admin", tags=["import-export"], dependencies=[Depends(require_role("admin"))])

STUDENT_HEADERS = ["الاسم الكامل", "البريد الإلكتروني", "الجامعة", "المرحلة"]
QUESTION_HEADERS = ["المادة", "نص السؤال", "الخيار 1", "الخيار 2", "الخيار 3", "الخيار 4", "رقم الإجابة الصحيحة (1-4)", "الشرح", "العنوان الفرعي"]


def _xlsx_response(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # Content-Disposition headers must be Latin-1 — an Arabic filename needs
    # the RFC 5987 filename* form, with a plain ASCII name as fallback.
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"export.xlsx\"; filename*=UTF-8''{quote(filename)}"},
    )


# ---------------------------------------------------------------- templates
@router.get("/import/template/students")
def students_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "طلاب"
    ws.append(STUDENT_HEADERS)
    ws.append(["مثال: زهراء علي", "zahra.example@uob.edu.iq", "جامعة بغداد", "المرحلة الثانية"])
    return _xlsx_response(wb, "قالب طلاب.xlsx")


@router.get("/import/template/questions")
def questions_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "أسئلة"
    ws.append(QUESTION_HEADERS)
    ws.append([
        "التشريح", "مثال: أي عظم أطول عظم في جسم الإنسان؟",
        "عظم الفخذ", "عظم العضد", "عظم الترقوة", "عظم الساق",
        1, "عظم الفخذ هو أطول وأقوى عظم في جسم الإنسان.", "عام",
    ])
    return _xlsx_response(wb, "قالب أسئلة.xlsx")


# ------------------------------------------------------------------- export
@router.get("/export")
def export_data(datasets: str, db: Session = Depends(get_db)):
    """datasets: comma-separated subset of students,questions,orders,logs"""
    wanted = {d.strip() for d in datasets.split(",") if d.strip()}
    wb = Workbook()
    wb.remove(wb.active)

    if "students" in wanted:
        ws = wb.create_sheet("الطلاب")
        ws.append(["الاسم", "البريد الإلكتروني", "الجامعة", "المرحلة", "محظور"])
        for s in db.query(models.User).filter(models.User.role == models.Role.student).all():
            ws.append([
                s.full_name, s.email,
                s.university.name if s.university else "",
                s.stage.name if s.stage else "",
                "نعم" if s.is_banned else "لا",
            ])

    if "questions" in wanted:
        ws = wb.create_sheet("الأسئلة")
        ws.append(["المادة", "نص السؤال", "الخيارات", "الإجابة الصحيحة"])
        for q in db.query(models.Question).all():
            choices_text = " | ".join(c.text for c in q.choices)
            correct = next((c.text for c in q.choices if c.is_correct), "")
            ws.append([q.subject.name if q.subject else "", q.text, choices_text, correct])

    if "orders" in wanted:
        ws = wb.create_sheet("الطلبات")
        ws.append(["رقم الطلب", "المستخدم", "الإجمالي", "طريقة الدفع", "الحالة", "التاريخ"])
        for o in db.query(models.Order).all():
            user = db.get(models.User, o.user_id)
            ws.append([
                o.id, user.full_name if user else "", o.total, o.payment_method, o.status,
                o.created_at.strftime("%Y-%m-%d %H:%M") if o.created_at else "",
            ])

    if "logs" in wanted:
        ws = wb.create_sheet("سجل النشاطات")
        ws.append(["الوقت", "المستخدم", "الإجراء", "IP"])
        logs = db.query(models.ActivityLog).order_by(models.ActivityLog.created_at.desc()).limit(500).all()
        for entry in logs:
            user = db.get(models.User, entry.user_id) if entry.user_id else None
            ws.append([
                entry.created_at.strftime("%Y-%m-%d %H:%M") if entry.created_at else "",
                user.full_name if user else "", entry.action, entry.ip_address,
            ])

    if not wb.sheetnames:
        ws = wb.create_sheet("بيانات")
        ws.append(["لم يتم تحديد أي بيانات للتصدير"])

    return _xlsx_response(wb, "نبض-تصدير.xlsx")


# ------------------------------------------------------------------- import
def _import_students(db: Session, rows) -> dict:
    created, skipped, errors = 0, 0, []
    for i, row in enumerate(rows, start=2):
        if not row or not any(row):
            continue
        padded = list(row) + [None] * 4
        name, email, uni_name, stage_name = padded[:4]
        if not email or not name:
            errors.append(f"صف {i}: الاسم أو البريد الإلكتروني ناقص")
            continue
        email = str(email).strip()
        if db.query(models.User).filter(models.User.email == email).first():
            skipped += 1
            continue
        university = (
            db.query(models.University).filter(models.University.name == str(uni_name).strip()).first()
            if uni_name else None
        )
        stage = (
            db.query(models.Stage).filter(models.Stage.name == str(stage_name).strip()).first()
            if stage_name else None
        )
        db.add(models.User(
            email=email, full_name=str(name).strip(), role=models.Role.student,
            university_id=university.id if university else None,
            stage_id=stage.id if stage else None,
        ))
        created += 1
    db.commit()
    return {"kind": "students", "created": created, "skipped": skipped, "errors": errors}


def _import_questions(db: Session, rows) -> dict:
    created, errors = 0, []
    for i, row in enumerate(rows, start=2):
        if not row or not any(row):
            continue
        padded = list(row) + [None] * 9
        subject_name, text, c1, c2, c3, c4, correct_idx, rationale, eyebrow = padded[:9]
        if not subject_name or not text or not c1 or not c2:
            errors.append(f"صف {i}: بيانات ناقصة (المادة، نص السؤال، وخيارين على الأقل مطلوبة)")
            continue
        subject = db.query(models.Subject).filter(models.Subject.name == str(subject_name).strip()).first()
        if not subject:
            errors.append(f'صف {i}: المادة "{subject_name}" غير موجودة بالمنهج')
            continue
        try:
            correct_idx = int(correct_idx) if correct_idx else 1
        except (TypeError, ValueError):
            correct_idx = 1
        choices = [c for c in [c1, c2, c3, c4] if c]
        q = models.Question(
            subject_id=subject.id, text=str(text).strip(),
            rationale=str(rationale).strip() if rationale else "",
            eyebrow=str(eyebrow).strip() if eyebrow else "",
        )
        db.add(q)
        db.flush()
        for idx, choice_text in enumerate(choices, start=1):
            db.add(models.Choice(
                question_id=q.id, text=str(choice_text).strip(),
                is_correct=(idx == correct_idx), order_index=idx - 1,
            ))
        created += 1
    db.commit()
    return {"kind": "questions", "created": created, "errors": errors}


@router.post("/import")
async def import_data(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Single dropzone for both templates — detects which one was uploaded
    from its header row rather than making the admin pick a file type."""
    wb = load_workbook(io.BytesIO(await file.read()), read_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    header = [str(h).strip() if h else "" for h in (header_row or [])]

    if header[:2] == STUDENT_HEADERS[:2]:
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        return _import_students(db, rows)
    if header[:2] == QUESTION_HEADERS[:2]:
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        return _import_questions(db, rows)
    raise HTTPException(400, "تنسيق الملف غير معروف — استخدمي أحد القوالب المتاحة للتنزيل")
